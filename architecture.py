import os
import sys
import time
import uuid
import json
import math
import threading
import warnings
import ctypes
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, ROUND_DOWN, InvalidOperation

import requests
from openpyxl import Workbook, load_workbook

import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.scrolledtext import ScrolledText

try:
    from urllib3.exceptions import InsecureRequestWarning
except Exception:
    InsecureRequestWarning = Warning


BASE_URL = "https://invest-public-api.tbank.ru/rest/tinkoff.public.invest.api.contract.v1"

APP_DIR = Path(__file__).resolve().parent
JOURNAL_PATH = APP_DIR / "trades_diary.xlsx"
STATE_PATH = APP_DIR / "active_trades.json"
ENV_FILE_USED: Path | None = None

# Dark palette close to the terminal screenshot.
BG = "#0F1A24"
BG_ALT = "#132231"
PANEL_BG = "#172838"
PANEL_BG_2 = "#1B2E3F"
HEADER_BG = "#0B141D"
INPUT_BG = "#223648"
INPUT_FG = "#F2EDE3"
FG = "#F2EDE3"
MUTED_FG = "#AEB8C2"
GREEN = "#3DDC97"
RED = "#FF5B6A"
YELLOW = "#F2C94C"
BLUE = "#5DADEC"
BORDER = "#31495F"
BUTTON_BG = "#102033"
BUTTON_ACTIVE = "#1E3A55"
BUY_BUTTON_BG = "#079B65"
BUY_BUTTON_ACTIVE = "#0BB978"
SELL_BUTTON_BG = "#A62A2A"
SELL_BUTTON_ACTIVE = "#C03535"
TREE_SELECTED = "#244A67"
FONT_FAMILY = "Calibri"

DEFAULT_RR = Decimal("2")
RISK_PER_TRADE_PERCENT = Decimal("1")
RISK_PER_TRADE_FRACTION = RISK_PER_TRADE_PERCENT / Decimal("100")
ENTRY_WAIT_SECONDS = 12
OCO_CHECK_SECONDS = 3

SSL_VERIFY = False
warnings.simplefilter("ignore", InsecureRequestWarning)


def load_local_env() -> None:
    global ENV_FILE_USED
    candidates = [APP_DIR / ".env", APP_DIR / ".evn"]
    env_path = next((p for p in candidates if p.exists()), None)
    if not env_path:
        return

    ENV_FILE_USED = env_path
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        os.environ[key] = value


load_local_env()
TOKEN = os.getenv("TINVEST_TOKEN") or os.getenv("INVEST_TOKEN")


def die(message: str):
    raise RuntimeError(message)


def headers() -> dict:
    if not TOKEN:
        die("Не задан TINVEST_TOKEN. Создай .env рядом со скриптом и добавь: TINVEST_TOKEN=твой_токен")
    return {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json",
    }


def post(method: str, payload: dict) -> dict:
    url = f"{BASE_URL}.{method}"
    response = requests.post(url, headers=headers(), json=payload, timeout=25, verify=SSL_VERIFY)
    if response.status_code >= 400:
        raise RuntimeError(f"API {method} вернул {response.status_code}: {response.text}")
    return response.json()


def q_to_decimal(q: dict | None) -> Decimal:
    if not q:
        return Decimal("0")
    units = Decimal(str(q.get("units", "0") or "0"))
    nano = Decimal(str(q.get("nano", 0) or 0)) / Decimal("1000000000")
    return units + nano


def decimal_to_q(value: Decimal) -> dict:
    value = Decimal(value)
    sign = -1 if value < 0 else 1
    value_abs = abs(value)
    units = int(value_abs)
    nano = int((value_abs - Decimal(units)) * Decimal("1000000000"))
    if sign < 0:
        units = -units
        nano = -nano
    return {"units": str(units), "nano": nano}


def money_to_decimal(m: dict | None) -> Decimal:
    return q_to_decimal(m)


def api_value_to_decimal(value) -> Decimal:
    if isinstance(value, dict):
        return money_to_decimal(value)
    if isinstance(value, (int, float, str, Decimal)):
        try:
            return Decimal(str(value))
        except Exception:
            return Decimal("0")
    return Decimal("0")


def get_nested_value(data: dict, *names):
    for name in names:
        if isinstance(data, dict) and name in data and data.get(name) not in (None, ""):
            return data.get(name)
    return None


def get_lot_size(instrument: dict, full: dict) -> Decimal:
    raw = get_nested_value(full, "lot") or get_nested_value(instrument, "lot") or 1
    try:
        lot = Decimal(str(raw))
    except Exception:
        lot = Decimal("1")
    return lot if lot > 0 else Decimal("1")


def get_price_point_value(instrument: dict, full: dict, step: Decimal) -> Decimal:
    # For futures, T-Invest can provide minPriceIncrementAmount: money value of one minimum price step.
    # Therefore money value of 1 full price point = minPriceIncrementAmount / minPriceIncrement.
    raw_tick_money = (
        get_nested_value(full, "minPriceIncrementAmount", "min_price_increment_amount")
        or get_nested_value(instrument, "minPriceIncrementAmount", "min_price_increment_amount")
    )
    tick_money = api_value_to_decimal(raw_tick_money)
    if tick_money > 0 and step > 0:
        return tick_money / step

    # Fallback for shares/currencies/bonds: risk per 1 price unit is price distance * lot size.
    return get_lot_size(instrument, full)


def calc_sl_by_risk(entry: Decimal, side: str, step: Decimal, qty: int, risk_amount: Decimal, point_value: Decimal) -> tuple[Decimal, Decimal]:
    if qty <= 0:
        die("Количество должно быть больше нуля.")
    if risk_amount <= 0:
        die("Не вижу положительный риск-бюджет для расчёта SL.")
    if point_value <= 0:
        die("Не смог определить стоимость пункта инструмента.")

    raw_distance = risk_amount / (Decimal(qty) * point_value)
    if raw_distance <= 0:
        die("SL дистанция получилась нулевой.")

    if side == "BUY":
        raw_sl = entry - raw_distance
        sl = round_to_step(raw_sl, step)
        if step > 0 and sl >= entry:
            sl = entry - step
        if sl <= 0:
            die(f"Авто-SL для LONG получился ниже/равен нулю: entry={entry}, distance={raw_distance}.")
    elif side == "SELL":
        raw_sl = entry + raw_distance
        sl = round_to_step(raw_sl, step)
        if step > 0 and sl <= entry:
            sl = entry + step
    else:
        die("side должен быть BUY или SELL")

    actual_distance = abs(entry - sl)
    if actual_distance <= 0:
        die("SL совпал с ценой входа после округления. Уменьши количество или проверь шаг цены.")
    return sl, actual_distance


def extract_int_from_keys(data, keys: list[str]) -> int | None:
    """Find an integer value by key, including nested dict/list API responses."""
    if data is None:
        return None

    wanted = {key.lower() for key in keys}

    def walk(node):
        if isinstance(node, dict):
            for key, value in node.items():
                if str(key).lower() in wanted and value not in (None, ""):
                    try:
                        return int(value)
                    except Exception:
                        pass
            for value in node.values():
                found = walk(value)
                if found is not None:
                    return found
        elif isinstance(node, list):
            for item in node:
                found = walk(item)
                if found is not None:
                    return found
        return None

    return walk(data)


def get_available_lots_from_max_lots(data: dict, side: str) -> int | None:
    if not isinstance(data, dict):
        return None

    if side == "BUY":
        views = [
            data.get("buyLimits"),
            data.get("buy_limits"),
            data.get("buyMarginLimits"),
            data.get("buy_margin_limits"),
            data,
        ]
        keys = [
            "buyMaxLots", "buy_max_lots",
            "buyMaxMarketLots", "buy_max_market_lots",
            "buyMaxLimitLots", "buy_max_limit_lots",
            "maxBuyLots", "max_buy_lots",
            "maxBuyMarketLots", "max_buy_market_lots",
            "maxLots", "max_lots",
        ]
    else:
        # For opening shorts, own-position sell limits can be zero. Margin limits are the useful capacity.
        views = [
            data.get("sellMarginLimits"),
            data.get("sell_margin_limits"),
            data.get("shortLimits"),
            data.get("short_limits"),
            data.get("sellLimits"),
            data.get("sell_limits"),
            data,
        ]
        keys = [
            "sellMaxLots", "sell_max_lots",
            "sellMaxMarketLots", "sell_max_market_lots",
            "sellMaxLimitLots", "sell_max_limit_lots",
            "maxSellLots", "max_sell_lots",
            "maxSellMarketLots", "max_sell_market_lots",
            "shortMaxLots", "short_max_lots",
            "maxShortLots", "max_short_lots",
            "maxLots", "max_lots",
        ]

    found = []
    for view in views:
        value = extract_int_from_keys(view, keys)
        if value is not None:
            found.append(value)
    return max(found) if found else None


def is_auto_sl_value(value: str) -> bool:
    return str(value).strip().lower() in {"auto", "авто"}


def is_percent_sl_value(value: str) -> bool:
    return str(value).strip().replace(",", ".").endswith("%")


def parse_sl_risk_percent(value: str) -> Decimal:
    raw = str(value).strip().replace(",", ".")
    if is_percent_sl_value(raw):
        number = raw[:-1].strip()
        if not number:
            die("В процентах SL нет числа. Пример: 0.5% или 2%")
        try:
            percent = Decimal(number)
        except (InvalidOperation, ValueError):
            die("SL в процентах должен быть числом. Пример: 0.5% или 2%")
        if percent <= 0:
            die("SL в процентах должен быть больше нуля.")
        return percent
    return RISK_PER_TRADE_PERCENT


def is_risk_based_sl_value(value: str) -> bool:
    return is_auto_sl_value(value) or is_percent_sl_value(value)


def calc_position_risk(distance: Decimal, qty: int, point_value: Decimal) -> Decimal:
    return abs(distance) * Decimal(qty) * point_value


def calc_position_value(entry: Decimal, qty: int, point_value: Decimal) -> Decimal:
    return abs(entry) * Decimal(qty) * point_value


def calc_position_based_risk_budget(entry: Decimal, qty: int, point_value: Decimal) -> Decimal:
    return calc_position_value(entry, qty, point_value) * RISK_PER_TRADE_FRACTION


def round_to_step(price: Decimal, step: Decimal) -> Decimal:
    if step <= 0:
        return price
    return (price / step).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * step


def parse_decimal(raw: str, field_name: str) -> Decimal:
    value = str(raw).strip().replace(",", ".")
    if not value:
        die(f"Поле «{field_name}» пустое.")
    try:
        result = Decimal(value)
    except (InvalidOperation, ValueError):
        die(f"Поле «{field_name}» должно быть числом. Пример: 110 или 109.85")
    if result <= 0:
        die(f"Поле «{field_name}» должно быть больше нуля.")
    return result


def fmt_dec(value, places: int = 2) -> str:
    try:
        d = Decimal(str(value))
        return f"{d.quantize(Decimal('1.' + '0' * places))}"
    except Exception:
        return str(value)


def fmt_money(value) -> str:
    try:
        d = Decimal(str(value))
        return f"{d.quantize(Decimal('1.00'))} ₽"
    except Exception:
        return str(value)


def fmt_percent(value) -> str:
    try:
        d = Decimal(str(value))
        return f"{d.quantize(Decimal('1.00'))}%"
    except Exception:
        return "—"


def signed_text(value: Decimal, suffix: str = " ₽") -> str:
    try:
        sign = "+" if value > 0 else ""
        return f"{sign}{fmt_dec(value)}{suffix}"
    except Exception:
        return "—"


def price_type_for(class_code: str) -> str:
    if str(class_code).upper() == "SPBFUT":
        return "PRICE_TYPE_POINT"
    return "PRICE_TYPE_CURRENCY"


def side_to_text(side: str) -> str:
    return "LONG" if side == "BUY" else "SHORT"


def position_side_from_qty(qty: Decimal) -> str:
    if qty > 0:
        return "BUY"
    if qty < 0:
        return "SELL"
    return "FLAT"


# ---------------------------- API helpers ----------------------------

def get_accounts() -> list[dict]:
    return post("UsersService/GetAccounts", {}).get("accounts", [])


def get_portfolio(account_id: str) -> dict:
    return post("OperationsService/GetPortfolio", {"accountId": account_id})


def get_positions(account_id: str) -> dict:
    return post("OperationsService/GetPositions", {"accountId": account_id})


def get_withdraw_limits(account_id: str) -> dict:
    return post("OperationsService/GetWithdrawLimits", {"accountId": account_id})


def get_max_lots(account_id: str, instrument_id: str, price: Decimal | None = None) -> dict:
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
    }
    if price is not None and price > 0:
        payload["price"] = decimal_to_q(price)
    return post("OrdersService/GetMaxLots", payload)


def get_stop_orders(account_id: str) -> list[dict]:
    return post("StopOrdersService/GetStopOrders", {"accountId": account_id}).get("stopOrders", [])


def find_instrument(ticker: str) -> dict:
    raw = ticker.strip().upper()
    preferred_class_code = None
    search_ticker = raw

    if "_" in raw:
        search_ticker, preferred_class_code = raw.split("_", 1)
        search_ticker = search_ticker.strip().upper()
        preferred_class_code = preferred_class_code.strip().upper()

    data = post("InstrumentsService/FindInstrument", {
        "query": search_ticker,
        "apiTradeAvailableFlag": True,
    })

    exact = []
    for inst in data.get("instruments", []):
        ticker_match = str(inst.get("ticker", "")).upper() == search_ticker
        class_match = not preferred_class_code or str(inst.get("classCode", "")).upper() == preferred_class_code
        if ticker_match and class_match:
            exact.append(inst)

    if not exact:
        die(
            f"Не нашёл торговый инструмент по тикеру {ticker}. "
            f"Для фьючерса можно вводить TICKER_CLASSCODE, например BRM6_SPBFUT."
        )

    exact.sort(key=lambda x: 0 if str(x.get("classCode", "")).upper() == "SPBFUT" else 1)
    return exact[0]


def get_instrument_id(inst: dict) -> str:
    instrument_id = inst.get("uid") or inst.get("instrumentUid") or inst.get("instrumentId") or inst.get("figi")
    if not instrument_id:
        die(f"У инструмента нет uid/figi. Ответ API: {inst}")
    return instrument_id


def get_instrument_full_by_uid(uid: str) -> dict:
    if not uid:
        return {}
    data = post("InstrumentsService/GetInstrumentBy", {
        "idType": "INSTRUMENT_ID_TYPE_UID",
        "id": uid,
    })
    return data.get("instrument", {})


def get_instrument_full_by_figi(figi: str) -> dict:
    if not figi:
        return {}
    data = post("InstrumentsService/GetInstrumentBy", {
        "idType": "INSTRUMENT_ID_TYPE_FIGI",
        "id": figi,
    })
    return data.get("instrument", {})


def get_min_step(instrument: dict, full: dict) -> Decimal:
    raw = full.get("minPriceIncrement") or instrument.get("minPriceIncrement")
    if raw:
        return q_to_decimal(raw)
    return Decimal("0.01")


def get_best_entry_price(instrument_id: str, side: str) -> Decimal:
    data = post("MarketDataService/GetOrderBook", {
        "instrumentId": instrument_id,
        "depth": 1,
    })
    bids = data.get("bids", [])
    asks = data.get("asks", [])

    if side == "BUY":
        if not asks:
            die("В стакане нет ask. Нельзя открыть LONG по текущей лимитной цене.")
        return q_to_decimal(asks[0]["price"])

    if side == "SELL":
        if not bids:
            die("В стакане нет bid. Нельзя открыть SHORT по текущей лимитной цене.")
        return q_to_decimal(bids[0]["price"])

    die("side должен быть BUY или SELL")


def get_order_book(
    instrument_id: str | None = None,
    depth: int = 50,
    ticker: str | None = None,
    instrument_uid: str | None = None,
    figi: str | None = None,
) -> dict:
    """Return raw T-Invest order book for Predictor/Analysis wall detection.

    The function accepts either a resolved instrument id/uid/figi or a ticker.
    Predi uses this wrapper to scan bid/ask walls without knowing the exact
    broker API call shape.
    """
    resolved_id = str(instrument_id or instrument_uid or figi or "").strip()

    # If the caller passes a plain ticker, resolve it to a T-Invest instrument id.
    if not resolved_id and ticker:
        inst = find_instrument(str(ticker).strip().upper())
        resolved_id = get_instrument_id(inst)

    if not resolved_id:
        die("Не передан instrument_id/ticker для получения стакана.")

    try:
        depth_int = int(depth or 50)
    except Exception:
        depth_int = 50
    depth_int = max(1, min(depth_int, 50))

    return post("MarketDataService/GetOrderBook", {
        "instrumentId": resolved_id,
        "depth": depth_int,
    })


def get_orderbook(*args, **kwargs) -> dict:
    return get_order_book(*args, **kwargs)


def load_order_book(*args, **kwargs) -> dict:
    return get_order_book(*args, **kwargs)


def fetch_order_book(*args, **kwargs) -> dict:
    return get_order_book(*args, **kwargs)


def calc_tp_sl(entry: Decimal, side: str, step: Decimal, sl_price: Decimal, tp_price: Decimal | None = None) -> tuple[Decimal, Decimal]:
    sl = round_to_step(sl_price, step)

    if side == "BUY":
        if sl >= entry:
            die(f"Для LONG стоп должен быть ниже входа. Вход: {entry}, SL: {sl}")
        if tp_price is None:
            tp = entry + ((entry - sl) * DEFAULT_RR)
        else:
            tp = tp_price
            if tp <= entry:
                die(f"Для LONG TP должен быть выше входа. Вход: {entry}, TP: {tp}")
    elif side == "SELL":
        if sl <= entry:
            die(f"Для SHORT стоп должен быть выше входа. Вход: {entry}, SL: {sl}")
        if tp_price is None:
            tp = entry - ((sl - entry) * DEFAULT_RR)
        else:
            tp = tp_price
            if tp >= entry:
                die(f"Для SHORT TP должен быть ниже входа. Вход: {entry}, TP: {tp}")
    else:
        die("side должен быть BUY или SELL")

    return round_to_step(tp, step), sl


def calc_sl_optional_tp(entry: Decimal, side: str, step: Decimal, sl_price: Decimal, tp_price: Decimal | None = None) -> tuple[Decimal | None, Decimal]:
    """Validate and round SL. If TP field is empty, TP is intentionally not created."""
    sl = round_to_step(sl_price, step)
    tp = round_to_step(tp_price, step) if tp_price is not None else None

    if side == "BUY":
        if sl >= entry:
            die(f"Для LONG стоп должен быть ниже входа. Вход: {entry}, SL: {sl}")
        if tp is not None and tp <= entry:
            die(f"Для LONG TP должен быть выше входа. Вход: {entry}, TP: {tp}")
    elif side == "SELL":
        if sl <= entry:
            die(f"Для SHORT стоп должен быть выше входа. Вход: {entry}, SL: {sl}")
        if tp is not None and tp >= entry:
            die(f"Для SHORT TP должен быть ниже входа. Вход: {entry}, TP: {tp}")
    else:
        die("side должен быть BUY или SELL")

    return tp, sl


def fmt_optional_price(value: Decimal | None) -> str:
    return str(value) if value is not None else "—"


def post_limit_entry(account_id: str, instrument_id: str, qty: int, price: Decimal, side: str, class_code: str) -> str:
    direction = "ORDER_DIRECTION_BUY" if side == "BUY" else "ORDER_DIRECTION_SELL"
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
        "quantity": str(qty),
        "price": decimal_to_q(price),
        "direction": direction,
        "orderType": "ORDER_TYPE_LIMIT",
        "orderId": str(uuid.uuid4()),
        "timeInForce": "TIME_IN_FORCE_FILL_AND_KILL",
        "priceType": price_type_for(class_code),
        "confirmMarginTrade": side == "SELL",
    }
    data = post("OrdersService/PostOrder", payload)
    order_id = data.get("orderId")
    if not order_id:
        die(f"Не получил orderId от API. Ответ: {data}")
    return order_id


def post_market_order(account_id: str, instrument_id: str, qty: int, direction: str, class_code: str, confirm_margin: bool = False) -> str:
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
        "quantity": str(qty),
        "price": decimal_to_q(Decimal("0")),
        "direction": direction,
        "orderType": "ORDER_TYPE_MARKET",
        "orderId": str(uuid.uuid4()),
        "timeInForce": "TIME_IN_FORCE_DAY",
        "priceType": price_type_for(class_code),
        "confirmMarginTrade": confirm_margin,
    }
    data = post("OrdersService/PostOrder", payload)
    order_id = data.get("orderId")
    if not order_id:
        die(f"Не получил orderId от API. Ответ: {data}")
    return order_id


def get_order_state(account_id: str, order_id: str) -> dict:
    return post("OrdersService/GetOrderState", {
        "accountId": account_id,
        "orderId": order_id,
    })


def wait_fill(account_id: str, order_id: str, timeout_sec: int = ENTRY_WAIT_SECONDS) -> int:
    last_state = None
    for _ in range(timeout_sec):
        state = get_order_state(account_id, order_id)
        last_state = state
        status = str(state.get("executionReportStatus", "")).upper()
        lots_executed = int(state.get("lotsExecuted", 0) or 0)
        if "FILL" in status and lots_executed > 0:
            return lots_executed
        if "REJECT" in status or "CANCEL" in status:
            return lots_executed
        time.sleep(1)
    return int((last_state or {}).get("lotsExecuted", 0) or 0)


def post_stop(account_id: str, instrument_id: str, qty: int, price: Decimal, original_side: str, stop_type: str, class_code: str) -> str:
    exit_direction = "STOP_ORDER_DIRECTION_SELL" if original_side == "BUY" else "STOP_ORDER_DIRECTION_BUY"
    payload = {
        "accountId": account_id,
        "instrumentId": instrument_id,
        "quantity": str(qty),
        "price": decimal_to_q(price),
        "stopPrice": decimal_to_q(price),
        "direction": exit_direction,
        "expirationType": "STOP_ORDER_EXPIRATION_TYPE_GOOD_TILL_CANCEL",
        "stopOrderType": stop_type,
        "exchangeOrderType": "EXCHANGE_ORDER_TYPE_MARKET",
        "priceType": price_type_for(class_code),
        "orderId": str(uuid.uuid4()),
        "confirmMarginTrade": False,
    }
    data = post("StopOrdersService/PostStopOrder", payload)
    stop_id = data.get("stopOrderId")
    if not stop_id:
        die(f"Не получил stopOrderId. Ответ: {data}")
    return stop_id


def cancel_stop_order(account_id: str, stop_order_id: str):
    if not stop_order_id:
        return
    post("StopOrdersService/CancelStopOrder", {
        "accountId": account_id,
        "stopOrderId": stop_order_id,
    })


# ---------------------------- State and journal ----------------------------

def load_state() -> dict:
    if not STATE_PATH.exists():
        return {"active_trades": []}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"active_trades": []}


def save_state(state: dict):
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def add_active_trade(trade: dict):
    state = load_state()
    state.setdefault("active_trades", []).append(trade)
    save_state(state)


def update_active_trade(trade_id: str, updates: dict):
    state = load_state()
    for trade in state.get("active_trades", []):
        if trade.get("trade_id") == trade_id:
            trade.update(updates)
            break
    save_state(state)


def remove_active_trade(trade_id: str):
    state = load_state()
    state["active_trades"] = [t for t in state.get("active_trades", []) if t.get("trade_id") != trade_id]
    save_state(state)


def find_active_trade(account_id: str, ticker: str | None = None, side: str | None = None, instrument_id: str | None = None) -> dict | None:
    state = load_state()
    ticker_upper = ticker.upper().strip() if ticker else None
    for trade in reversed(state.get("active_trades", [])):
        if trade.get("account_id") != account_id:
            continue
        if instrument_id and trade.get("instrument_id") != instrument_id:
            continue
        if ticker_upper and str(trade.get("ticker", "")).upper() != ticker_upper:
            continue
        if side and trade.get("side") != side:
            continue
        return trade
    return None


def ensure_journal():
    headers_row = ["Дата", "Капитал счёта", "PnL", "Тикер", "Комментарии"]
    if JOURNAL_PATH.exists():
        wb = load_workbook(JOURNAL_PATH)
        ws = wb.active
        current = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if current[:5] != headers_row:
            ws = wb.create_sheet("Trades")
            ws.append(headers_row)
        wb.save(JOURNAL_PATH)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(headers_row)
    widths = [14, 18, 14, 16, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    wb.save(JOURNAL_PATH)


def append_journal_row(date_str: str, capital: Decimal | None, pnl: Decimal | None, ticker: str, comments: str = ""):
    ensure_journal()
    wb = load_workbook(JOURNAL_PATH)
    ws = wb.active
    if [cell.value for cell in ws[1]][:5] != ["Дата", "Капитал счёта", "PnL", "Тикер", "Комментарии"]:
        ws = wb["Trades"] if "Trades" in wb.sheetnames else wb.create_sheet("Trades")
        if ws.max_row == 1 and not ws[1][0].value:
            ws.append(["Дата", "Капитал счёта", "PnL", "Тикер", "Комментарии"])

    ws.append([
        date_str,
        float(capital) if capital is not None else None,
        float(pnl) if pnl is not None else None,
        ticker,
        comments,
    ])
    wb.save(JOURNAL_PATH)


def get_total_portfolio_value(account_id: str) -> Decimal:
    return money_to_decimal(get_portfolio(account_id).get("totalAmountPortfolio"))


def get_total_selected_portfolio_value(account_ids: list[str]) -> Decimal:
    total = Decimal("0")
    for account_id in account_ids:
        total += get_total_portfolio_value(account_id)
    return total


# ---------------------------- Chart helpers ----------------------------

def get_candles(instrument_id: str, interval: str = "CANDLE_INTERVAL_1_MIN", minutes: int = 90) -> list[dict]:
    """Load recent candles for a chart tile. Keeps the payload small for low-latency UI updates."""
    from datetime import timedelta

    minutes = max(5, int(minutes or 90))
    to_dt = datetime.utcnow().replace(microsecond=0)
    from_dt = to_dt - timedelta(minutes=minutes)
    data = post("MarketDataService/GetCandles", {
        "instrumentId": instrument_id,
        "from": from_dt.isoformat() + "Z",
        "to": to_dt.isoformat() + "Z",
        "interval": interval,
    })
    return data.get("candles", []) or []


def candle_to_ohlc(candle: dict) -> dict:
    """Normalize a T-Invest candle into Decimal OHLC values used by chart_panel.py."""
    return {
        "time": candle.get("time") or candle.get("timestamp") or "",
        "open": q_to_decimal(candle.get("open")),
        "high": q_to_decimal(candle.get("high")),
        "low": q_to_decimal(candle.get("low")),
        "close": q_to_decimal(candle.get("close")),
        "volume": int(candle.get("volume", 0) or 0),
    }




def get_chart_live_price(instrument_id: str) -> Decimal:
    """Fast lightweight price read for chart live updates using top-of-book midpoint."""
    data = post("MarketDataService/GetOrderBook", {
        "instrumentId": instrument_id,
        "depth": 1,
    })
    bids = data.get("bids", []) or []
    asks = data.get("asks", []) or []

    bid = q_to_decimal(bids[0].get("price")) if bids else Decimal("0")
    ask = q_to_decimal(asks[0].get("price")) if asks else Decimal("0")

    if bid > 0 and ask > 0:
        return (bid + ask) / Decimal("2")
    if ask > 0:
        return ask
    if bid > 0:
        return bid
    return Decimal("0")


def get_chart_live_prices(instrument_ids: list[str]) -> dict[str, Decimal]:
    """Batch live prices for chart tiles. Falls back to top-of-book midpoint reads."""
    ids = []
    for instrument_id in instrument_ids or []:
        if instrument_id and instrument_id not in ids:
            ids.append(instrument_id)
    if not ids:
        return {}

    try:
        try:
            data = post("MarketDataService/GetLastPrices", {"instrumentId": ids})
        except Exception:
            data = post("MarketDataService/GetLastPrices", {"instrumentIds": ids})

        raw_prices = data.get("lastPrices") or data.get("last_prices") or []
        result: dict[str, Decimal] = {}
        unresolved = []

        for item in raw_prices:
            if not isinstance(item, dict):
                continue
            price = q_to_decimal(item.get("price"))
            if price <= 0:
                continue
            item_id = (
                item.get("instrumentUid")
                or item.get("instrument_uid")
                or item.get("instrumentId")
                or item.get("instrument_id")
                or item.get("uid")
                or item.get("figi")
            )
            if item_id:
                result[str(item_id)] = price
            else:
                unresolved.append(price)

        if unresolved and len(unresolved) == len(ids):
            for instrument_id, price in zip(ids, unresolved):
                result.setdefault(instrument_id, price)

        if len(raw_prices) == len(ids):
            for instrument_id, item in zip(ids, raw_prices):
                if instrument_id in result or not isinstance(item, dict):
                    continue
                price = q_to_decimal(item.get("price"))
                if price > 0:
                    result[instrument_id] = price

        if result:
            return result
    except Exception:
        pass

    result: dict[str, Decimal] = {}
    for instrument_id in ids:
        try:
            price = get_chart_live_price(instrument_id)
            if price > 0:
                result[instrument_id] = price
        except Exception:
            continue
    return result

def load_chart_candles_by_ticker(ticker: str, interval: str = "CANDLE_INTERVAL_1_MIN", minutes: int = 90) -> dict:
    """Resolve ticker and return compact candle data for the chart tab."""
    instrument = find_instrument(ticker)
    instrument_id = get_instrument_id(instrument)
    full = get_instrument_full_by_uid(instrument_id)
    candles = [candle_to_ohlc(c) for c in get_candles(instrument_id, interval, minutes)]
    candles = [c for c in candles if c["close"] > 0]
    return {
        "ticker": str(instrument.get("ticker") or ticker).upper(),
        "instrument_id": instrument_id,
        "class_code": instrument.get("classCode") or full.get("classCode") or "",
        "name": instrument.get("name") or full.get("name") or "",
        "candles": candles,
    }
